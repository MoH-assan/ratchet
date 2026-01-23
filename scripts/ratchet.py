import argparse
import sys
from pathlib import Path
from typing import Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

try:
    from .helper import (
        ErrorLog,
        SHEET_PRESTEMP,
        SHEET_PROPERTIES,
        calculate_allowable,
        coerce_numeric,
        compute_envelope,
        errors_to_dataframe,
        extract_properties,
        extract_runner_columns,
        find_sheet_name,
        log_error,
        normalize_columns,
        normalize_pipe_id,
        parse_cases,
        read_excel_sheet,
    )
except ImportError:
    # Allow running this file directly: python scripts/ratchet.py
    sys.path.append(str(Path(__file__).resolve().parents[1]))
    from scripts.helper import (  # type: ignore
        ErrorLog,
        SHEET_PRESTEMP,
        SHEET_PROPERTIES,
        calculate_allowable,
        coerce_numeric,
        compute_envelope,
        errors_to_dataframe,
        extract_properties,
        extract_runner_columns,
        find_sheet_name,
        log_error,
        normalize_columns,
        normalize_pipe_id,
        parse_cases,
        read_excel_sheet,
    )


def _series_max(series: pd.Series, *, use_abs: bool = False) -> float:
    values = pd.to_numeric(series, errors="coerce")
    if use_abs:
        values = values.abs()
    if values.dropna().empty:
        return float("nan")
    return float(values.max())


def _series_min(series: pd.Series) -> float:
    values = pd.to_numeric(series, errors="coerce")
    if values.dropna().empty:
        return float("nan")
    return float(values.min())


def _from_to_for_extreme(
    df: pd.DataFrame,
    value_col: str,
    *,
    mode: str,
    use_abs: bool = False,
) -> Tuple[Optional[str], Optional[str]]:
    if value_col not in df.columns or "from" not in df.columns or "to" not in df.columns:
        return None, None
    values = pd.to_numeric(df[value_col], errors="coerce")
    if use_abs:
        values = values.abs()
    if values.dropna().empty:
        return None, None
    idx = values.idxmax() if mode == "max" else values.idxmin()
    from_val = df.loc[idx, "from"]
    to_val = df.loc[idx, "to"]
    if pd.isna(from_val) or pd.isna(to_val):
        return None, None
    return str(from_val), str(to_val)


def _from_to_for_max_abs(df: pd.DataFrame, value_col: str) -> Tuple[Optional[str], Optional[str]]:
    return _from_to_for_extreme(df, value_col, mode="max", use_abs=True)


def _from_to_for_max(df: pd.DataFrame, value_col: str) -> Tuple[Optional[str], Optional[str]]:
    return _from_to_for_extreme(df, value_col, mode="max", use_abs=False)


def _from_to_for_min(df: pd.DataFrame, value_col: str) -> Tuple[Optional[str], Optional[str]]:
    return _from_to_for_extreme(df, value_col, mode="min", use_abs=False)


def _format_from_to(from_val: Optional[str], to_val: Optional[str]) -> Optional[str]:
    if not from_val or not to_val:
        return None
    return f"{from_val}->{to_val}"


def build_material_envelopes(summary: pd.DataFrame) -> Optional[pd.DataFrame]:
    if "pipe_material" not in summary.columns:
        return None
    material_series = summary["pipe_material"].dropna().astype(str).str.strip()
    unique_materials = [m for m in material_series.unique() if m]
    if not unique_materials:
        return None

    extra_cols = []
    for col in ("d_out_case", "thck_case", "c4_case"):
        if col not in summary.columns:
            extra_cols.append(col)

    all_columns = list(summary.columns) + extra_cols
    rows = []
    for material in unique_materials:
        material_norm = str(material).strip()
        subset = summary.loc[
            summary["pipe_material"].astype(str).str.strip() == material_norm
        ]
        if subset.empty:
            continue
        row = {col: np.nan for col in all_columns}
        if "material" in row:
            row["material"] = material_norm
        row["pipe_material"] = material_norm

        if "p_max" in row:
            row["p_max"] = _series_max(subset["p_max"], use_abs=True)
        if "sy_min" in row:
            row["sy_min"] = _series_min(subset["sy_min"])
        if "delta_t1_max" in row:
            row["delta_t1_max"] = _series_max(subset["delta_t1_max"])
        if "E_max" in row:
            row["E_max"] = _series_max(subset["E_max"])
        if "d_out" in row:
            row["d_out"] = _series_max(subset["d_out"])
        if "thck" in row:
            row["thck"] = _series_min(subset["thck"])
        if "alpha_room" in row:
            row["alpha_room"] = _series_max(subset["alpha_room"])
        if "c4" in row:
            row["c4"] = _series_min(subset["c4"])

        if "p_max_case" in row:
            row["p_max_case"] = _format_from_to(*_from_to_for_max_abs(subset, "p_max"))
        if "sy_min_case" in row:
            row["sy_min_case"] = _format_from_to(*_from_to_for_min(subset, "sy_min"))
        if "delta_t1_case" in row:
            row["delta_t1_case"] = _format_from_to(*_from_to_for_max(subset, "delta_t1_max"))
        if "E_max_case" in row:
            row["E_max_case"] = _format_from_to(*_from_to_for_max(subset, "E_max"))
        if "d_out_case" in row:
            row["d_out_case"] = _format_from_to(*_from_to_for_max(subset, "d_out"))
        if "thck_case" in row:
            row["thck_case"] = _format_from_to(*_from_to_for_min(subset, "thck"))
        if "c4_case" in row:
            row["c4_case"] = _format_from_to(*_from_to_for_min(subset, "c4"))

        allowable, note, x_val, y_val = calculate_allowable(pd.Series(row))
        if "allowable" in row:
            row["allowable"] = allowable
        if "allowable_note" in row:
            row["allowable_note"] = note
        if "x" in row:
            row["x"] = x_val
        if "y" in row:
            row["y"] = y_val

        rows.append(row)

    if not rows:
        return None
    return pd.DataFrame(rows, columns=all_columns)


def apply_output_formatting(output_path: Path) -> None:
    wb = load_workbook(output_path)
    node_units_map = {
        "from": "",
        "to": "",
        "material": "",
        "pipe_id": "",
        "nominal_in": "in",
        "p_max": "psi",
        "p_max_case": "cntr load case #",
        "sy_min": "psi",
        "sy_min_case": "cntr load case #",
        "delta_t1_max": "deg F",
        "delta_t1_case": "cntr load case #",
        "e_max": "E6 psi",
        "e_max_case": "cntr load case #",
        "d_out": "in",
        "thck": "in",
        "pipe_material": "",
        "alpha_room": "E-6 in/inF",
        "c4": "",
        "allowable": "deg F",
        "allowable_note": "",
        "x": "",
        "y": "",
    }

    material_units_map = dict(node_units_map)
    for key in (
        "p_max_case",
        "sy_min_case",
        "delta_t1_case",
        "e_max_case",
        "d_out_case",
        "thck_case",
        "c4_case",
    ):
        material_units_map[key] = "cntr runner f2"

    def _apply_units(ws: object, units_map: dict) -> dict:
        headers = [cell.value for cell in ws[1]]
        header_map = {
            str(value).strip().lower(): idx + 1
            for idx, value in enumerate(headers)
            if value is not None
        }
        ws.insert_rows(2)
        for idx, header in enumerate(headers, start=1):
            key = str(header).strip().lower()
            units = units_map.get(key, "")
            ws.cell(row=2, column=idx, value=units)
        return header_map

    def _cell_text(ws: object, row_idx: int, col_idx: int) -> str:
        value = ws.cell(row=row_idx, column=col_idx).value
        return "" if value is None else str(value).strip()

    def _split_from_to(value: object) -> Tuple[Optional[str], Optional[str]]:
        if value is None:
            return None, None
        text = str(value)
        if "->" not in text:
            return None, None
        left, right = text.split("->", 1)
        left = left.strip()
        right = right.strip()
        if not left or not right:
            return None, None
        return left, right

    ws_node = wb["PerNodeEnvlope"] if "PerNodeEnvlope" in wb.sheetnames else None
    ws_material = wb["PerMaterialEnvlope"] if "PerMaterialEnvlope" in wb.sheetnames else None

    node_header_map = _apply_units(ws_node, node_units_map) if ws_node else {}
    material_header_map = _apply_units(ws_material, material_units_map) if ws_material else {}

    def _hsv_to_rgb(hue: float, sat: float, val: float) -> Tuple[int, int, int]:
        if sat == 0:
            rgb = int(val * 255)
            return rgb, rgb, rgb
        hue = hue % 1.0
        h_i = int(hue * 6.0)
        f = hue * 6.0 - h_i
        p = val * (1.0 - sat)
        q = val * (1.0 - sat * f)
        t = val * (1.0 - sat * (1.0 - f))
        h_i = h_i % 6
        if h_i == 0:
            r, g, b = val, t, p
        elif h_i == 1:
            r, g, b = q, val, p
        elif h_i == 2:
            r, g, b = p, val, t
        elif h_i == 3:
            r, g, b = p, q, val
        elif h_i == 4:
            r, g, b = t, p, val
        else:
            r, g, b = val, p, q
        return int(r * 255), int(g * 255), int(b * 255)

    def _color_for_index(idx: int) -> str:
        golden_ratio = 0.61803398875
        hue = (idx * golden_ratio) % 1.0
        r, g, b = _hsv_to_rgb(hue, 0.75, 0.85)
        return f"FF{r:02X}{g:02X}{b:02X}"

    material_colors: dict[str, str] = {}
    if ws_material and material_header_map.get("pipe_material"):
        material_col = material_header_map["pipe_material"]
        for row_idx in range(3, ws_material.max_row + 1):
            material_val = _cell_text(ws_material, row_idx, material_col)
            if not material_val:
                continue
            if material_val not in material_colors:
                material_colors[material_val] = _color_for_index(len(material_colors))

    def _color_for_material(material_val: Optional[str]) -> str:
        if not material_val:
            return "FFD62728"
        return material_colors.get(material_val, "FFD62728")

    envelope_row = None
    if ws_material and material_header_map:
        from_col = material_header_map.get("from")
        to_col = material_header_map.get("to")
        if from_col and to_col:
            for row_idx in range(3, ws_material.max_row + 1):
                from_val = ws_material.cell(row=row_idx, column=from_col).value
                to_val = ws_material.cell(row=row_idx, column=to_col).value
                if str(from_val).strip().lower() == "feeder" and str(to_val).strip().lower() == "envloped":
                    envelope_row = row_idx
                    break

    if ws_material and material_header_map:
        normal_font = Font(bold=False, color="FF000000")
        calc_keys = ("x", "y", "allowable", "allowable_note")
        case_keys = ("p_max_case", "sy_min_case", "delta_t1_case", "e_max_case", "d_out_case", "thck_case", "c4_case")
        for row_idx in range(3, ws_material.max_row + 1):
            material_val = None
            if material_header_map.get("pipe_material"):
                material_val = _cell_text(ws_material, row_idx, material_header_map["pipe_material"])
            color = _color_for_material(material_val)
            bold_font = Font(bold=True, color=color)
            for col_idx in range(1, ws_material.max_column + 1):
                ws_material.cell(row=row_idx, column=col_idx).font = bold_font
            for key in case_keys:
                col_idx = material_header_map.get(key)
                if col_idx:
                    ws_material.cell(row=row_idx, column=col_idx).font = normal_font
            for key in calc_keys:
                col_idx = material_header_map.get(key)
                if col_idx:
                    ws_material.cell(row=row_idx, column=col_idx).font = normal_font

    if ws_node and node_header_map and material_header_map and ws_material:
        from_col = node_header_map.get("from")
        to_col = node_header_map.get("to")
        pipe_material_col = node_header_map.get("pipe_material")
        if from_col and to_col:
            case_to_value = {
                "p_max_case": "p_max",
                "sy_min_case": "sy_min",
                "delta_t1_case": "delta_t1_max",
                "e_max_case": "e_max",
                "d_out_case": "d_out",
                "thck_case": "thck",
                "c4_case": "c4",
            }
            for mat_row in range(3, ws_material.max_row + 1):
                material_value = None
                if material_header_map.get("pipe_material"):
                    material_value = _cell_text(ws_material, mat_row, material_header_map["pipe_material"])
                bold_font = Font(bold=True, color=_color_for_material(material_value))
                for case_key, value_key in case_to_value.items():
                    case_col = material_header_map.get(case_key)
                    value_col = node_header_map.get(value_key)
                    if not case_col or not value_col:
                        continue
                    from_val, to_val = _split_from_to(
                        ws_material.cell(row=mat_row, column=case_col).value
                    )
                    if not from_val or not to_val:
                        continue
                    target_row = None
                    for row_idx in range(3, ws_node.max_row + 1):
                        if _cell_text(ws_node, row_idx, from_col).lower() != from_val.lower():
                            continue
                        if _cell_text(ws_node, row_idx, to_col).lower() != to_val.lower():
                            continue
                        if pipe_material_col and material_value:
                            material = _cell_text(ws_node, row_idx, pipe_material_col).upper()
                            if material != material_value.upper():
                                continue
                        target_row = row_idx
                        break
                    if target_row:
                        ws_node.cell(row=target_row, column=value_col).font = bold_font

    wb.save(output_path)
    wb.close()


def reorder_material_columns(df: pd.DataFrame) -> pd.DataFrame:
    cols = [col for col in df.columns if col not in ("from", "to")]
    for base, case in (("d_out", "d_out_case"), ("thck", "thck_case"), ("c4", "c4_case")):
        if base in cols and case in cols:
            cols.remove(case)
            insert_at = cols.index(base) + 1
            cols.insert(insert_at, case)
    return df.loc[:, cols]


def process_file(file_path: Path, output_dir: Path) -> Tuple[Optional[Path], ErrorLog]:
    errors: ErrorLog = ErrorLog()
    file_name = file_path.name

    try:
        xls = pd.ExcelFile(file_path, engine="openpyxl")
    except Exception as exc:
        log_error(errors, f"Failed to open Excel file: {exc}", file_name=file_name)
        return None, errors

    pres_sheet = find_sheet_name(xls.sheet_names, SHEET_PRESTEMP)
    prop_sheet = find_sheet_name(xls.sheet_names, SHEET_PROPERTIES)

    if not pres_sheet or not prop_sheet:
        log_error(
            errors,
            "Missing required sheet(s).",
            file_name=file_name,
        )
        print(
            f"[{file_name}] Expected sheets: '{SHEET_PRESTEMP}', '{SHEET_PROPERTIES}'. "
            f"Found: {xls.sheet_names}"
        )
        return None, errors

    pres_df = read_excel_sheet(file_path, pres_sheet)
    prop_df = read_excel_sheet(file_path, prop_sheet)

    pres_df = normalize_columns(pres_df, errors, file_name, pres_sheet)
    prop_df = normalize_columns(prop_df, errors, file_name, prop_sheet)

    pres_df["row_id"] = pres_df.index

    base_df, runner_cols = extract_runner_columns(pres_df, errors, file_name, pres_sheet)
    base_df["row_id"] = pres_df["row_id"]

    long_df, case_numbers = parse_cases(pres_df, runner_cols, errors, file_name, pres_sheet)

    if long_df.empty:
        output_path = output_dir / f"{file_path.stem}_ratchet.xlsx"
        output_dir.mkdir(parents=True, exist_ok=True)
        errors_df = errors_to_dataframe(errors)
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            pd.DataFrame().to_excel(writer, sheet_name="RatchetSummary", index=False)
            errors_df.to_excel(writer, sheet_name="Errors", index=False)
        return output_path, errors

    numeric_cols = [
        col
        for col in long_df.columns
        if col not in runner_cols + ["row_id", "case_number"]
    ]
    long_df = coerce_numeric(
        long_df,
        numeric_cols,
        errors,
        file_name,
        pres_sheet,
    )

    envelope_df = compute_envelope(long_df, errors, file_name, pres_sheet)

    prop_subset = extract_properties(prop_df, errors, file_name, prop_sheet)
    prop_subset = coerce_numeric(
        prop_subset,
        ["d_out", "thck", "alpha_room", "c4"],
        errors,
        file_name,
        prop_sheet,
    )

    if "pipe_id" in prop_subset.columns:
        prop_subset["pipe_id_norm"] = prop_subset["pipe_id"].apply(normalize_pipe_id)
        duplicate_ids = prop_subset["pipe_id_norm"].duplicated(keep=False)
        if duplicate_ids.any():
            for pid in sorted(set(prop_subset.loc[duplicate_ids, "pipe_id_norm"].dropna())):
                log_error(
                    errors,
                    f"Multiple PipeProperties rows found for Pipe ID '{pid}'. Using first.",
                    file_name=file_name,
                    sheet=prop_sheet,
                    level="warning",
                )
            prop_subset = prop_subset.drop_duplicates("pipe_id_norm", keep="first")
    else:
        prop_subset["pipe_id_norm"] = None

    summary = base_df.merge(envelope_df, on="row_id", how="left")
    if "pipe_id" in summary.columns:
        summary["pipe_id_norm"] = summary["pipe_id"].apply(normalize_pipe_id)
    else:
        summary["pipe_id_norm"] = None

    summary = summary.merge(prop_subset, on="pipe_id_norm", how="left", suffixes=("", "_prop"))

    for idx, row in summary.iterrows():
        if pd.isna(row.get("pipe_id_norm")):
            log_error(
                errors,
                "Missing Pipe ID for join.",
                file_name=file_name,
                sheet=pres_sheet,
                row=int(row.get("row_id", idx)) + 3,
                level="warning",
            )
            continue
        if pd.isna(row.get("d_out")):
            log_error(
                errors,
                f"No PipeProperties match for Pipe ID '{row.get('pipe_id_norm')}'.",
                file_name=file_name,
                sheet=prop_sheet,
                level="warning",
            )

    allowables = summary.apply(calculate_allowable, axis=1)
    summary["allowable"] = allowables.apply(lambda item: item[0])
    summary["allowable_note"] = allowables.apply(lambda item: item[1])
    summary["x"] = allowables.apply(lambda item: item[2])
    summary["y"] = allowables.apply(lambda item: item[3])

    envelope_df = build_material_envelopes(summary)

    summary = summary.drop(columns=["row_id", "pipe_id_norm"], errors="ignore")

    output_path = output_dir / f"{file_path.stem}_ratchet.xlsx"
    output_dir.mkdir(parents=True, exist_ok=True)

    errors_df = errors_to_dataframe(errors)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="PerNodeEnvlope", index=False)
        if envelope_df is None:
            pd.DataFrame(columns=summary.columns).to_excel(
                writer, sheet_name="PerMaterialEnvlope", index=False
            )
        else:
            envelope_df = reorder_material_columns(envelope_df)
            envelope_df.to_excel(writer, sheet_name="PerMaterialEnvlope", index=False)
        errors_df.to_excel(writer, sheet_name="Errors", index=False)

    apply_output_formatting(output_path)

    print(
        f"[{file_name}] rows={len(pres_df)}, cases={case_numbers}, "
        f"matches={summary['pipe_id'].notna().sum() if 'pipe_id' in summary.columns else 0}, "
        f"output={output_path}"
    )

    return output_path, errors


def collect_input_files(input_dir: Path) -> list[Path]:
    return [
        path
        for path in sorted(input_dir.glob("*.xlsx"))
        if not path.name.startswith("~$")
    ]


def main() -> int:
    parser = argparse.ArgumentParser(description="Ratchet automation runner")
    parser.add_argument(
        "--input",
        dest="input_dir",
        type=str,
        default=None,
        help="Input directory containing .xlsx files",
    )
    parser.add_argument(
        "--output",
        dest="output_dir",
        type=str,
        default=None,
        help="Output directory for results",
    )
    args = parser.parse_args()

    base_dir = Path(__file__).resolve().parents[1]
    input_dir = Path(args.input_dir) if args.input_dir else base_dir / "data" / "input"
    output_dir = Path(args.output_dir) if args.output_dir else base_dir / "data" / "output"

    files = collect_input_files(input_dir)
    if not files:
        print(f"No input .xlsx files found in {input_dir}")
        return 1

    for file_path in files:
        try:
            process_file(file_path, output_dir)
        except Exception as exc:
            print(f"[{file_path.name}] Failed: {exc}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
